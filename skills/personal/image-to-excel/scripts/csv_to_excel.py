#!/usr/bin/env python3
"""
CSV转Excel工具 - 带格式美化

用法:
    python3 csv_to_excel.py <csv文件路径> [输出xlsx路径]

示例:
    python3 csv_to_excel.py data.csv
    python3 csv_to_excel.py data.csv output.xlsx
"""

import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要安装 openpyxl")
    print("运行: pip install openpyxl")
    sys.exit(1)


def csv_to_excel(csv_path: str, xlsx_path: str = None, sheet_name: str = "数据表") -> str:
    """
    将CSV文件转换为带格式的Excel文件

    Args:
        csv_path: CSV文件路径
        xlsx_path: 输出Excel文件路径（可选，默认与CSV同名）
        sheet_name: 工作表名称

    Returns:
        生成的Excel文件路径
    """
    csv_path = Path(csv_path)

    if not csv_path.exists():
        raise FileNotFoundError(f"CSV文件不存在: {csv_path}")

    # 默认输出路径
    if xlsx_path is None:
        xlsx_path = csv_path.with_suffix('.xlsx')
    else:
        xlsx_path = Path(xlsx_path)

    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # 定义样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 读取CSV并写入Excel
    col_widths = {}  # 记录每列最大宽度

    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

                # 标题行样式
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment

                # 计算列宽（中文字符算2个宽度）
                str_value = str(value)
                width = sum(2 if ord(c) > 127 else 1 for c in str_value)
                col_widths[col_idx] = max(col_widths.get(col_idx, 10), width + 2)

    # 设置列宽
    for col_idx, width in col_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(width, 50)  # 最大50

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存
    wb.save(xlsx_path)
    print(f"Excel文件已生成: {xlsx_path}")

    return str(xlsx_path)


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    csv_path = sys.argv[1]
    xlsx_path = sys.argv[2] if len(sys.argv) > 2 else None

    try:
        result = csv_to_excel(csv_path, xlsx_path)
        print(f"转换成功: {result}")
    except Exception as e:
        print(f"转换失败: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
