#!/usr/bin/env python3
"""
赔偿计算器 - 劳动/工伤/交通/离婚赔偿

用法:
    python3 calc_compensation.py --type labor --monthly-salary 8000 --years 3.5 --reason illegal --output 赔偿.xlsx
    python3 calc_compensation.py --type injury --monthly-salary 6000 --years 5 --disability-level 8 --output 工伤.xlsx
    python3 calc_compensation.py --type traffic --urban --disability-level 10 --lost-work-days 30 --nursing-days 15 --hospital-days 10 --medical-fee 25000 --output 交通.xlsx
    python3 calc_compensation.py --type divorce --assets-json 财产.json --output 财产分割.xlsx

参数说明:
    --type          计算类型: labor/injury/traffic/divorce
    --monthly-salary  月工资（元）
    --years         工龄（年，支持小数）
    --reason        解除原因: illegal（违法）/ legal（合法）/ mutual（协商）
    --disability-level  伤残等级（1-10级，1级最重）
    --urban         使用城镇标准（不加则用农村标准）
    --lost-work-days  误工天数
    --nursing-days  护理天数
    --hospital-days 住院天数
    --medical-fee   医疗费用（元）
    --assets-json   共同财产清单 JSON 文件路径（离婚用）
    --output        输出 Excel 文件路径
    --params-json   标准参数 JSON 文件路径（默认使用内置参数）
"""

import argparse
import json
import math
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误: 需要安装 openpyxl")
    print("运行: pip install openpyxl")
    sys.exit(1)


# ===== 默认标准参数（用户可通过 references/标准参数.json 更新）=====
DEFAULT_PARAMS = {
    "year": 2024,
    "national_avg_salary": 124110,      # 全国城镇非私营单位年均工资（元），2024年数据
    "urban_avg_annual": 54188,          # 城镇居民人均可支配收入（元/年），2024年数据
    "rural_avg_annual": 23119,          # 农村居民人均可支配收入（元/年），2024年数据
    "urban_avg_daily": 148.46,          # 城镇日均（元/天）= 54188 / 365
    "rural_avg_daily": 63.34,           # 农村日均（元/天）= 23119 / 365
    "life_expectancy": 79,              # 人均寿命（岁），2024年数据
    "upper_limit_multiplier": 3,        # 经济补偿金月工资上限倍数
    "upper_limit_cap_years": 12,        # 经济补偿金最高年限
    "note": "以上数据仅供参考，请根据当年最新数据更新 references/标准参数.json"
}

# 工伤伤残补助金标准（月数，按1-10级）
INJURY_DISABILITY_MONTHS = {
    1: 27, 2: 25, 3: 23, 4: 21, 5: 18,
    6: 16, 7: 13, 8: 11, 9: 8, 10: 6
}

# 工伤医疗补助金月数（5-10级，解除合同时）
INJURY_MEDICAL_MONTHS = {5: 22, 6: 18, 7: 14, 8: 10, 9: 8, 10: 6}

# 工伤就业补助金月数（5-10级，解除合同时）
INJURY_EMPLOYMENT_MONTHS = {5: 36, 6: 30, 7: 24, 8: 18, 9: 12, 10: 6}

# 交通事故伤残赔偿系数（按1-10级）
TRAFFIC_DISABILITY_RATE = {
    1: 1.0, 2: 0.9, 3: 0.8, 4: 0.7, 5: 0.6,
    6: 0.5, 7: 0.4, 8: 0.3, 9: 0.2, 10: 0.1
}


def load_params(params_json_path=None):
    """加载标准参数，优先使用外部文件"""
    if params_json_path and Path(params_json_path).exists():
        with open(params_json_path, 'r', encoding='utf-8') as f:
            params = json.load(f)
        return {**DEFAULT_PARAMS, **params}
    # 尝试加载同目录下的标准参数文件
    default_path = Path(__file__).parent.parent / 'references' / '标准参数.json'
    if default_path.exists():
        with open(default_path, 'r', encoding='utf-8') as f:
            params = json.load(f)
        return {**DEFAULT_PARAMS, **params}
    return DEFAULT_PARAMS


def calc_labor(monthly_salary, years, reason, params):
    """
    计算劳动争议经济补偿金/赔偿金
    根据《劳动合同法》第46、47条
    """
    # 月工资上限（当地月平均工资3倍）
    local_avg_monthly = params['national_avg_salary'] / 12
    salary_cap = local_avg_monthly * params['upper_limit_multiplier']
    effective_salary = min(monthly_salary, salary_cap)

    # 计算工龄月数（不满半年按0.5计，满半年不满一年按1计）
    if years - int(years) < 0.5:
        n = int(years) + (0.5 if years - int(years) > 0 else 0)
    else:
        n = math.ceil(years)
    n = min(n, params['upper_limit_cap_years'])  # 上限12年

    compensation = effective_salary * n

    items = [
        ["计算项目", "计算过程", "金额（元）", "法律依据"],
        ["月工资基数", f"实际月工资 ¥{monthly_salary:.2f}", f"{monthly_salary:.2f}",
         "《劳动合同法》第47条"],
        ["月工资上限", f"当地月均工资 ¥{local_avg_monthly:.2f} × {params['upper_limit_multiplier']}倍",
         f"{salary_cap:.2f}", "《劳动合同法》第47条第2款"],
        ["实际计算工资", f"取月工资与上限较小值", f"{effective_salary:.2f}", ""],
        ["计算年限", f"工龄 {years} 年 → 计算系数 {n} 个月", f"N={n}", ""],
    ]

    if reason == 'illegal':
        items.append(["违法解除赔偿金", f"¥{effective_salary:.2f} × {n} × 2倍",
                       f"{compensation * 2:.2f}", "《劳动合同法》第87条"])
        result = compensation * 2
        remark = "违法解除劳动合同，支付2N赔偿金"
    else:
        items.append(["经济补偿金", f"¥{effective_salary:.2f} × {n}", f"{compensation:.2f}",
                       "《劳动合同法》第47条"])
        result = compensation
        remark = "合法解除或协商解除，支付N经济补偿金"

    summary = [
        ["汇总项目", "金额（元）"],
        ["经济补偿/赔偿金", f"{result:.2f}"],
        ["备注", remark],
        ["注意事项", f"月工资超过当地月均工资3倍({salary_cap:.2f}元)时，按上限计算，最多{params['upper_limit_cap_years']}年"],
    ]

    return items, summary, f"劳动争议经济{'赔偿' if reason == 'illegal' else '补偿'}金计算"


def calc_injury(monthly_salary, years, disability_level, params):
    """
    计算工伤赔偿
    根据《工伤保险条例》第35-37条
    """
    if disability_level not in range(1, 11):
        raise ValueError(f"伤残等级应为1-10级，当前输入：{disability_level}")

    # 一次性伤残补助金
    disability_months = INJURY_DISABILITY_MONTHS[disability_level]
    disability_allowance = monthly_salary * disability_months

    items = [
        ["计算项目", "计算过程", "金额（元）", "法律依据"],
        ["本人月工资", f"¥{monthly_salary:.2f}/月", f"{monthly_salary:.2f}",
         "《工伤保险条例》第35条"],
        ["伤残等级", f"{disability_level}级伤残", "-", ""],
        ["一次性伤残补助金",
         f"¥{monthly_salary:.2f} × {disability_months}个月",
         f"{disability_allowance:.2f}",
         f"《工伤保险条例》第{'35' if disability_level <= 4 else '36' if disability_level <= 6 else '37'}条"],
    ]

    total = disability_allowance

    # 5-10级：解除合同时还有医疗补助金和就业补助金
    if disability_level >= 5:
        medical_months = INJURY_MEDICAL_MONTHS.get(disability_level, 0)
        employment_months = INJURY_EMPLOYMENT_MONTHS.get(disability_level, 0)
        local_avg_monthly = params['national_avg_salary'] / 12

        medical_allowance = local_avg_monthly * medical_months
        employment_allowance = local_avg_monthly * employment_months

        items.extend([
            ["一次性医疗补助金（解除合同时）",
             f"当地月均工资¥{local_avg_monthly:.2f} × {medical_months}个月",
             f"{medical_allowance:.2f}",
             "《工伤保险条例》第37条"],
            ["一次性就业补助金（解除合同时）",
             f"当地月均工资¥{local_avg_monthly:.2f} × {employment_months}个月",
             f"{employment_allowance:.2f}",
             "《工伤保险条例》第37条"],
        ])
        total += medical_allowance + employment_allowance

    summary = [
        ["汇总项目", "金额（元）"],
        ["一次性伤残补助金", f"{disability_allowance:.2f}"],
    ]
    if disability_level >= 5:
        summary.extend([
            ["一次性医疗补助金", f"{medical_allowance:.2f}"],
            ["一次性就业补助金", f"{employment_allowance:.2f}"],
        ])
    summary.append(["合计", f"{total:.2f}"])
    summary.append(["注意", "1-4级工伤保留劳动关系，按月领取伤残津贴，不适用一次性解除"])

    return items, summary, f"{disability_level}级工伤赔偿金计算"


def calc_traffic(urban, disability_level, lost_work_days, nursing_days,
                  hospital_days, medical_fee, params, victim_age=35):
    """
    计算交通事故赔偿
    根据最高人民法院《关于审理人身损害赔偿案件适用法律若干问题的解释》
    """
    if urban:
        annual_income = params['urban_avg_annual']
        daily_rate = params['urban_avg_daily']
        residence_type = "城镇"
    else:
        annual_income = params['rural_avg_annual']
        daily_rate = params['rural_avg_daily']
        residence_type = "农村"

    # 残疾赔偿金：年收入 × 20年 × 伤残系数
    compensation_years = min(20, params['life_expectancy'] - victim_age)
    disability_rate = TRAFFIC_DISABILITY_RATE.get(disability_level, 0)
    disability_compensation = annual_income * compensation_years * disability_rate

    # 误工费
    lost_work_fee = daily_rate * lost_work_days

    # 护理费（按当地护工费标准，用日均代替）
    nursing_fee = daily_rate * nursing_days

    # 住院伙食补助费（参考标准：100元/天）
    hospital_meal = 100 * hospital_days

    # 精神损害抚慰金（根据伤残等级）
    mental_compensation_map = {
        1: 100000, 2: 90000, 3: 80000, 4: 70000, 5: 60000,
        6: 50000, 7: 40000, 8: 30000, 9: 20000, 10: 10000
    }
    mental_compensation = mental_compensation_map.get(disability_level, 0)

    items = [
        ["赔偿项目", "计算过程", "金额（元）", "法律依据"],
        ["居住地类型", residence_type, "-", ""],
        ["年人均收入", f"¥{annual_income:.2f}/年", f"{annual_income:.2f}", ""],
        ["日均收入", f"¥{daily_rate:.2f}/天", f"{daily_rate:.2f}", ""],
        ["伤残赔偿金",
         f"¥{annual_income:.2f} × {compensation_years}年 × {disability_rate*100:.0f}%",
         f"{disability_compensation:.2f}",
         "《人身损害赔偿司法解释》第12条"],
        ["医疗费用", "按实际发生", f"{medical_fee:.2f}", "第6条"],
        ["误工费", f"¥{daily_rate:.2f}/天 × {lost_work_days}天", f"{lost_work_fee:.2f}", "第7条"],
        ["护理费", f"¥{daily_rate:.2f}/天 × {nursing_days}天", f"{nursing_fee:.2f}", "第8条"],
        ["住院伙食补助费", f"100元/天 × {hospital_days}天", f"{hospital_meal:.2f}", "第10条"],
        ["精神损害抚慰金", f"{disability_level}级伤残参考标准", f"{mental_compensation:.2f}", "第18条"],
    ]

    total = (disability_compensation + medical_fee + lost_work_fee +
             nursing_fee + hospital_meal + mental_compensation)

    summary = [
        ["汇总项目", "金额（元）"],
        ["伤残赔偿金", f"{disability_compensation:.2f}"],
        ["医疗费用", f"{medical_fee:.2f}"],
        ["误工费", f"{lost_work_fee:.2f}"],
        ["护理费", f"{nursing_fee:.2f}"],
        ["住院伙食补助费", f"{hospital_meal:.2f}"],
        ["精神损害抚慰金", f"{mental_compensation:.2f}"],
        ["合计", f"{total:.2f}"],
        ["注意", "精神损害抚慰金、护理天数等实际以法院认定为准"],
    ]

    return items, summary, f"{disability_level}级交通事故赔偿金计算（{residence_type}标准）"


def calc_divorce(assets_json_path, params):
    """
    计算离婚财产分割参考方案
    """
    if not Path(assets_json_path).exists():
        raise FileNotFoundError(f"财产清单文件不存在: {assets_json_path}")

    with open(assets_json_path, 'r', encoding='utf-8') as f:
        assets_data = json.load(f)

    assets = assets_data.get('共同财产', [])
    total = sum(float(a.get('估值', 0)) for a in assets)
    half = total / 2

    items = [
        ["资产名称", "类型", "估值（元）", "建议归属", "补偿金额（元）", "备注"],
    ]

    party_a_total = 0
    for asset in assets:
        name = asset.get('名称', '')
        asset_type = asset.get('类型', '')
        value = float(asset.get('估值', 0))
        to_party = asset.get('建议归属', '协商')
        remark = asset.get('备注', '')

        if to_party == '甲方':
            party_a_total += value
            compensation = ''
        elif to_party == '乙方':
            compensation = ''
        else:
            compensation = ''

        items.append([name, asset_type, f"{value:.2f}", to_party, compensation, remark])

    party_b_total = total - party_a_total
    compensation_needed = abs(party_a_total - party_b_total) / 2

    summary = [
        ["汇总项目", "金额（元）"],
        ["共同财产总额", f"{total:.2f}"],
        ["各方应得份额（各50%）", f"{half:.2f}"],
        ["甲方分得财产", f"{party_a_total:.2f}"],
        ["乙方分得财产", f"{party_b_total:.2f}"],
        ["折价补偿金（多得方补偿少得方）", f"{compensation_needed:.2f}"],
        ["说明", "以上为平均分割参考方案，法院会综合考量照顾子女、照顾无过错方等因素"],
    ]

    return items, summary, "离婚共同财产分割参考方案"


def style_header_row(ws, row, fill_color="2F5496"):
    """设置表头行样式"""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", name="宋体")
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_data_row(ws, row, even=False):
    """设置数据行样式"""
    fill_color = "EBF1F8" if even else "FFFFFF"
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(name="宋体", size=10)
    border = Border(
        bottom=Side(style="thin", color="CCCCCC")
    )
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def create_excel(title, items, summary, output_path, params=None):
    """生成 Excel 文件（两个 Sheet）"""
    wb = openpyxl.Workbook()

    # Sheet1：分项明细
    ws1 = wb.active
    ws1.title = "分项明细"

    # 标题
    ws1.merge_cells(f"A1:{get_column_letter(len(items[0]))}1")
    title_cell = ws1['A1']
    title_cell.value = title
    title_cell.font = Font(bold=True, size=14, name="黑体")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # 生成日期
    ws1.merge_cells(f"A2:{get_column_letter(len(items[0]))}2")
    date_cell = ws1['A2']
    date_cell.value = f"生成日期：{datetime.now().strftime('%Y年%m月%d日')}"
    date_cell.font = Font(size=10, name="宋体", color="666666")
    date_cell.alignment = Alignment(horizontal="right")

    # 数据
    for row_idx, row_data in enumerate(items):
        ws_row = row_idx + 3
        for col_idx, value in enumerate(row_data, 1):
            ws1.cell(row=ws_row, column=col_idx, value=value)
        if row_idx == 0:
            style_header_row(ws1, ws_row)
        else:
            style_data_row(ws1, ws_row, even=(row_idx % 2 == 0))

    # 调整列宽
    for col in ws1.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col_letter].width = min(max(max_len * 2, 12), 50)

    # Sheet2：汇总合计
    ws2 = wb.create_sheet("汇总合计")
    ws2.merge_cells(f"A1:B1")
    ws2['A1'].value = "汇总合计"
    ws2['A1'].font = Font(bold=True, size=14, name="黑体")
    ws2['A1'].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 30

    for row_idx, row_data in enumerate(summary):
        ws_row = row_idx + 2
        for col_idx, value in enumerate(row_data, 1):
            ws2.cell(row=ws_row, column=col_idx, value=value)
        if row_idx == 0 and len(summary) > 0 and summary[0][0] in ["汇总项目", "资产名称"]:
            style_header_row(ws2, ws_row)
        else:
            style_data_row(ws2, ws_row, even=(row_idx % 2 == 0))

    # 汇总表列宽
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 20

    # Sheet3：法条声明
    ws3 = wb.create_sheet("法条声明")
    ws3.merge_cells("A1:B1")
    ws3['A1'].value = "法条可靠性声明"
    ws3['A1'].font = Font(bold=True, size=14, name="黑体")
    ws3['A1'].alignment = Alignment(horizontal="center")
    ws3.row_dimensions[1].height = 30

    _params = params or {}
    data_year = _params.get('year', '未知')
    disclaimer_rows = [
        ["项目", "说明"],
        ["参数数据来源", "国家统计局（数据年份：{}年）".format(data_year)],
        ["全国城镇非私营单位年均工资", "查询路径：国家统计局官网 → 年度数据 → 就业和工资"],
        ["城镇/农村居民人均可支配收入", "查询路径：国家统计局官网 → 年度数据 → 人民生活"],
        ["AI 生成内容提示",
         "法律依据中的条文引用由 AI 辅助生成，可能存在版本滞后或条文号偏差，仅供参考"],
        ["核实建议",
         "请登录国家法律法规数据库（flk.npc.gov.cn）或中国裁判文书网核实相关法条现行有效版本"],
        ["免责声明",
         "本计算结果仅为辅助参考，最终赔偿金额以法院判决或仲裁裁决为准"],
        ["生成日期", datetime.now().strftime("%Y年%m月%d日")],
    ]

    for row_idx, row_data in enumerate(disclaimer_rows):
        ws_row = row_idx + 2
        for col_idx, value in enumerate(row_data, 1):
            ws3.cell(row=ws_row, column=col_idx, value=value)
        if row_idx == 0:
            style_header_row(ws3, ws_row)
        else:
            style_data_row(ws3, ws_row, even=(row_idx % 2 == 0))

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 60

    wb.save(output_path)
    print(f"Excel 文件已生成: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="赔偿计算器 - 劳动/工伤/交通/离婚赔偿",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )
    parser.add_argument('--type', required=True, choices=['labor', 'injury', 'traffic', 'divorce'],
                        help='计算类型: labor/injury/traffic/divorce')
    parser.add_argument('--monthly-salary', type=float, help='月工资（元）')
    parser.add_argument('--years', type=float, help='工龄（年，支持小数）')
    parser.add_argument('--reason', choices=['illegal', 'legal', 'mutual'], default='legal',
                        help='解除原因: illegal（违法）/ legal（合法）/ mutual（协商）')
    parser.add_argument('--disability-level', type=int, choices=range(1, 11),
                        metavar='1-10', help='伤残等级（1-10级，1级最重）')
    parser.add_argument('--urban', action='store_true', help='城镇标准（不加则农村标准）')
    parser.add_argument('--lost-work-days', type=int, default=0, help='误工天数')
    parser.add_argument('--nursing-days', type=int, default=0, help='护理天数')
    parser.add_argument('--hospital-days', type=int, default=0, help='住院天数')
    parser.add_argument('--medical-fee', type=float, default=0, help='医疗费用（元）')
    parser.add_argument('--assets-json', help='共同财产清单 JSON 文件路径（离婚用）')
    parser.add_argument('--params-json', help='标准参数 JSON 文件路径')
    parser.add_argument('--output', default='赔偿计算结果.xlsx', help='输出 Excel 文件路径')

    args = parser.parse_args()
    params = load_params(args.params_json)

    try:
        if args.type == 'labor':
            if not args.monthly_salary or not args.years:
                parser.error("labor 类型需要 --monthly-salary 和 --years")
            items, summary, title = calc_labor(args.monthly_salary, args.years, args.reason, params)

        elif args.type == 'injury':
            if not args.monthly_salary or not args.years or not args.disability_level:
                parser.error("injury 类型需要 --monthly-salary、--years 和 --disability-level")
            items, summary, title = calc_injury(
                args.monthly_salary, args.years, args.disability_level, params)

        elif args.type == 'traffic':
            if not args.disability_level:
                parser.error("traffic 类型需要 --disability-level")
            items, summary, title = calc_traffic(
                args.urban, args.disability_level, args.lost_work_days,
                args.nursing_days, args.hospital_days, args.medical_fee, params)

        elif args.type == 'divorce':
            if not args.assets_json:
                parser.error("divorce 类型需要 --assets-json")
            items, summary, title = calc_divorce(args.assets_json, params)

        create_excel(title, items, summary, args.output, params)

    except Exception as e:
        print(f"计算失败: {e}")
        sys.exit(1)


if __name__ == '__main__':
    main()
