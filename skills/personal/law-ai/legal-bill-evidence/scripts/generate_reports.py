#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
账单流水分析报告生成脚本
输入: 标准化 JSON 数据文件 (wechat 或 bank_card 类型)
输出: Excel (.xlsx) + HTML (.html) + PDF (.pdf)

用法:
    python3 generate_reports.py --data wechat_data.json --output-dir ./分析报告/
    python3 generate_reports.py --data card5432_data.json --output-dir ./分析报告/ --name "分析报告-卡号5432"
"""

import argparse
import json
import re
import subprocess
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import quote

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ============================================================
# 通用样式
# ============================================================
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
MONEY_FORMAT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
YELLOW_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
LIGHT_RED = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
ROW_EXPENSE = PatternFill(start_color="FFF2F2", end_color="FFF2F2", fill_type="solid")
ROW_INCOME = PatternFill(start_color="F2FFF2", end_color="F2FFF2", fill_type="solid")


def style_header_row(ws, row, cols):
    """设置表头样式"""
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


# ============================================================
# 微信支付分析
# ============================================================
def analyze_wechat(data):
    """分析微信交易数据"""
    transactions = data['transactions']
    expense_txs = [t for t in transactions if t['direction'] == '支出']
    income_txs = [t for t in transactions if t['direction'] == '收入']
    other_txs = [t for t in transactions if t['direction'] == '其他']

    expense_total = sum(t['amount'] for t in expense_txs)
    income_total = sum(t['amount'] for t in income_txs)
    other_total = sum(t['amount'] for t in other_txs)

    # 按支付方式
    by_payment = defaultdict(lambda: {'count': 0, 'total': 0.0})
    for t in expense_txs:
        by_payment[t['payment']]['count'] += 1
        by_payment[t['payment']]['total'] += t['amount']

    # 按交易类型
    by_type = defaultdict(lambda: {'count': 0, 'total': 0.0})
    for t in expense_txs:
        by_type[t['type']]['count'] += 1
        by_type[t['type']]['total'] += t['amount']

    # 按月份
    by_month = defaultdict(lambda: {'expense': 0.0, 'income': 0.0, 'count': 0})
    for t in transactions:
        month = t['date'][:7]
        if t['direction'] == '支出':
            by_month[month]['expense'] += t['amount']
            by_month[month]['count'] += 1
        elif t['direction'] == '收入':
            by_month[month]['income'] += t['amount']

    # 大额支出
    large = sorted(expense_txs, key=lambda t: -t['amount'])[:30]

    # 金额分布
    ranges = {
        '0-10元': 0, '10-50元': 0, '50-100元': 0,
        '100-500元': 0, '500-1000元': 0, '1000-5000元': 0, '5000元以上': 0,
    }
    for t in expense_txs:
        a = t['amount']
        if a < 10: ranges['0-10元'] += 1
        elif a < 50: ranges['10-50元'] += 1
        elif a < 100: ranges['50-100元'] += 1
        elif a < 500: ranges['100-500元'] += 1
        elif a < 1000: ranges['500-1000元'] += 1
        elif a < 5000: ranges['1000-5000元'] += 1
        else: ranges['5000元以上'] += 1

    # 计算天数
    if transactions:
        dates = sorted(set(t['date'] for t in transactions))
        from datetime import date as dt_date
        d1 = dt_date.fromisoformat(dates[0])
        d2 = dt_date.fromisoformat(dates[-1])
        days = (d2 - d1).days + 1
    else:
        days = 1

    return {
        'transactions': transactions,
        'expense_txs': expense_txs,
        'income_txs': income_txs,
        'other_txs': other_txs,
        'expense_total': expense_total,
        'income_total': income_total,
        'other_total': other_total,
        'by_payment': dict(by_payment),
        'by_type': dict(by_type),
        'by_month': dict(sorted(by_month.items())),
        'large_expenses': large,
        'ranges': ranges,
        'days': days,
        'holder': data.get('holder', ''),
        'account': data.get('account', ''),
        'period': data.get('period', {}),
    }


# ============================================================
# 微信 Excel
# ============================================================
def create_wechat_excel(analyzed, filepath):
    """生成微信支付 Excel"""
    wb = Workbook()

    # Sheet1: 交易明细
    ws = wb.active
    ws.title = "交易明细"
    headers = ["日期", "时间", "交易类型", "收/支", "支付方式", "金额(元)", "交易对方"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    for i, t in enumerate(analyzed['transactions'], 2):
        ws.cell(row=i, column=1, value=t['date'])
        ws.cell(row=i, column=2, value=t['time'])
        ws.cell(row=i, column=3, value=t['type'])
        ws.cell(row=i, column=4, value=t['direction'])
        ws.cell(row=i, column=5, value=t['payment'])
        c = ws.cell(row=i, column=6, value=t['amount'])
        c.number_format = MONEY_FORMAT
        ws.cell(row=i, column=7, value=t['counterparty'])

        fill = ROW_EXPENSE if t['direction'] == '支出' else ROW_INCOME if t['direction'] == '收入' else None
        for col in range(1, 8):
            cell = ws.cell(row=i, column=col)
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 25

    # Sheet2: 汇总统计
    ws2 = wb.create_sheet("汇总统计")
    period = analyzed.get('period', {})
    period_str = f"{period.get('start', '')} 至 {period.get('end', '')}"

    ws2.cell(row=1, column=1, value="微信支付交易汇总统计")
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.cell(row=2, column=1, value=f"时间段: {period_str}")

    # 总体
    row = 4
    for h_col, h_val in enumerate(["项目", "笔数", "金额(元)"], 1):
        ws2.cell(row=row, column=h_col, value=h_val)
    style_header_row(ws2, row, 3)

    for label, count, total in [
        ("支出", len(analyzed['expense_txs']), analyzed['expense_total']),
        ("收入", len(analyzed['income_txs']), analyzed['income_total']),
        ("其他", len(analyzed['other_txs']), analyzed['other_total']),
    ]:
        row += 1
        ws2.cell(row=row, column=1, value=label).border = THIN_BORDER
        ws2.cell(row=row, column=2, value=count).border = THIN_BORDER
        c = ws2.cell(row=row, column=3, value=total)
        c.number_format = MONEY_FORMAT
        c.border = THIN_BORDER

    row += 1
    ws2.cell(row=row, column=1, value="净支出").font = Font(bold=True)
    ws2.cell(row=row, column=1).border = THIN_BORDER
    ws2.cell(row=row, column=2).border = THIN_BORDER
    c = ws2.cell(row=row, column=3, value=analyzed['expense_total'] - analyzed['income_total'])
    c.number_format = MONEY_FORMAT
    c.font = Font(bold=True, color="FF0000")
    c.border = THIN_BORDER

    # 按月统计
    row += 2
    for h_col, h_val in enumerate(["月份", "支出笔数", "支出金额", "收入金额"], 1):
        ws2.cell(row=row, column=h_col, value=h_val)
    style_header_row(ws2, row, 4)

    for month, info in analyzed['by_month'].items():
        row += 1
        ws2.cell(row=row, column=1, value=month).border = THIN_BORDER
        ws2.cell(row=row, column=2, value=info['count']).border = THIN_BORDER
        c = ws2.cell(row=row, column=3, value=info['expense'])
        c.number_format = MONEY_FORMAT
        c.border = THIN_BORDER
        c = ws2.cell(row=row, column=4, value=info['income'])
        c.number_format = MONEY_FORMAT
        c.border = THIN_BORDER

    # 按支付方式
    row += 2
    for h_col, h_val in enumerate(["支付方式", "笔数", "金额(元)"], 1):
        ws2.cell(row=row, column=h_col, value=h_val)
    style_header_row(ws2, row, 3)

    for method, info in sorted(analyzed['by_payment'].items(), key=lambda x: -x[1]['total']):
        row += 1
        ws2.cell(row=row, column=1, value=method).border = THIN_BORDER
        ws2.cell(row=row, column=2, value=info['count']).border = THIN_BORDER
        c = ws2.cell(row=row, column=3, value=info['total'])
        c.number_format = MONEY_FORMAT
        c.border = THIN_BORDER

    # 大额支出
    row += 2
    ws2.cell(row=row, column=1, value="大额支出(>=500元)")
    ws2.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    for h_col, h_val in enumerate(["序号", "日期", "金额(元)", "类型", "交易对方"], 1):
        ws2.cell(row=row, column=h_col, value=h_val)
    style_header_row(ws2, row, 5)

    for idx, t in enumerate(analyzed['large_expenses'], 1):
        if t['amount'] < 500:
            break
        row += 1
        ws2.cell(row=row, column=1, value=idx).border = THIN_BORDER
        ws2.cell(row=row, column=2, value=t['date']).border = THIN_BORDER
        c = ws2.cell(row=row, column=3, value=t['amount'])
        c.number_format = MONEY_FORMAT
        c.border = THIN_BORDER
        c.fill = LIGHT_RED
        ws2.cell(row=row, column=4, value=t['type']).border = THIN_BORDER
        ws2.cell(row=row, column=5, value=t['counterparty']).border = THIN_BORDER

    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 25

    wb.save(filepath)
    print(f"  Excel: {filepath}")


# ============================================================
# 银行卡 Excel
# ============================================================
def create_card_excel(data, filepath):
    """生成银行卡 Excel"""
    transactions = data['transactions']
    page_totals = data.get('page_totals')
    card_name = data.get('card_name', '银行卡')
    card_number = data.get('card_number', '')

    wb = Workbook()

    # Sheet1: 交易明细
    ws = wb.active
    ws.title = "交易明细"
    headers = ["日期", "金额(元)", "类型", "对方/说明"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    expense_total = 0
    income_total = 0

    for i, tx in enumerate(transactions, 2):
        date = tx['date']
        amount = tx['amount']
        tx_type = tx.get('type', '')
        desc = tx.get('description', '')

        ws.cell(row=i, column=1, value=date).border = THIN_BORDER
        c = ws.cell(row=i, column=2, value=amount)
        c.number_format = MONEY_FORMAT
        c.border = THIN_BORDER
        ws.cell(row=i, column=3, value=tx_type).border = THIN_BORDER
        ws.cell(row=i, column=4, value=desc).border = THIN_BORDER

        if amount < 0:
            expense_total += abs(amount)
            for col in range(1, 5):
                ws.cell(row=i, column=col).fill = ROW_EXPENSE
        else:
            income_total += amount
            for col in range(1, 5):
                ws.cell(row=i, column=col).fill = ROW_INCOME

    # 逐笔识别汇总
    row = len(transactions) + 3
    ws.cell(row=row, column=1, value="逐笔识别-支出合计").font = Font(bold=True)
    c = ws.cell(row=row, column=2, value=-expense_total)
    c.number_format = MONEY_FORMAT
    c.font = Font(bold=True, color="FF0000")

    row += 1
    ws.cell(row=row, column=1, value="逐笔识别-收入合计").font = Font(bold=True)
    c = ws.cell(row=row, column=2, value=income_total)
    c.number_format = MONEY_FORMAT
    c.font = Font(bold=True, color="008000")

    row += 1
    ws.cell(row=row, column=1, value="逐笔识别-净支出").font = Font(bold=True)
    c = ws.cell(row=row, column=2, value=income_total - expense_total)
    c.number_format = MONEY_FORMAT
    c.font = Font(bold=True, color="FF0000")

    # 银行清单权威合计
    if page_totals:
        row += 2
        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = ORANGE_FILL
            ws.cell(row=row, column=col).border = THIN_BORDER
        ws.cell(row=row, column=1, value="银行清单合计(准确值)").font = Font(bold=True, size=12)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

        row += 1
        ws.cell(row=row, column=1, value="银行清单-支出合计").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = YELLOW_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        c = ws.cell(row=row, column=2, value=-page_totals['expense'])
        c.number_format = MONEY_FORMAT
        c.font = Font(bold=True, color="FF0000", size=12)
        c.fill = YELLOW_FILL
        c.border = THIN_BORDER

        row += 1
        ws.cell(row=row, column=1, value="银行清单-收入合计").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = YELLOW_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        c = ws.cell(row=row, column=2, value=page_totals['income'])
        c.number_format = MONEY_FORMAT
        c.font = Font(bold=True, color="008000", size=12)
        c.fill = YELLOW_FILL
        c.border = THIN_BORDER

        row += 1
        ws.cell(row=row, column=1, value="银行清单-净支出").font = Font(bold=True)
        ws.cell(row=row, column=1).fill = YELLOW_FILL
        ws.cell(row=row, column=1).border = THIN_BORDER
        net = page_totals['income'] - page_totals['expense']
        c = ws.cell(row=row, column=2, value=net)
        c.number_format = MONEY_FORMAT
        c.font = Font(bold=True, color="FF0000", size=12)
        c.fill = YELLOW_FILL
        c.border = THIN_BORDER

        # 差异说明
        diff_exp = abs(expense_total - page_totals['expense'])
        row += 2
        if diff_exp > 0.5:
            note = (f"注: 逐笔识别支出={expense_total:,.2f}元, "
                    f"银行清单支出={page_totals['expense']:,.2f}元, "
                    f"差异={diff_exp:,.2f}元。"
                    f"差异原因: 银行清单为图片扫描件, 部分交易在识别时可能遗漏。"
                    f"以银行清单页脚合计为准。")
        else:
            note = "注: 逐笔识别合计与银行清单合计基本一致(差异<0.5元)。以银行清单页脚合计为准。"
        ws.cell(row=row, column=1, value=note).font = Font(italic=True, color="666666")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 35

    # Sheet2: 银行清单分页合计
    if page_totals and page_totals.get('pages'):
        ws2 = wb.create_sheet("银行清单分页合计")
        ws2.cell(row=1, column=1, value=f"{card_name} 银行清单分页汇总")
        ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
        holder = data.get('holder', '')
        branch = data.get('branch', '')
        ws2.cell(row=2, column=1, value=f"卡号: {card_number} | 户名: {holder} | {branch}")

        row = 4
        for h_col, h_val in enumerate(["页码", "支出(元)", "收入(元)"], 1):
            ws2.cell(row=row, column=h_col, value=h_val)
        style_header_row(ws2, row, 3)

        for page in page_totals['pages']:
            row += 1
            ws2.cell(row=row, column=1, value=page['name']).border = THIN_BORDER
            c = ws2.cell(row=row, column=2, value=page['expense'])
            c.number_format = MONEY_FORMAT
            c.border = THIN_BORDER
            c.font = Font(color="FF0000")
            c = ws2.cell(row=row, column=3, value=page['income'])
            c.number_format = MONEY_FORMAT
            c.border = THIN_BORDER
            c.font = Font(color="008000")

        row += 1
        ws2.cell(row=row, column=1, value="合计").font = Font(bold=True, size=12)
        ws2.cell(row=row, column=1).border = THIN_BORDER
        ws2.cell(row=row, column=1).fill = ORANGE_FILL
        c = ws2.cell(row=row, column=2, value=page_totals['expense'])
        c.number_format = MONEY_FORMAT
        c.font = Font(bold=True, color="FF0000", size=12)
        c.border = THIN_BORDER
        c.fill = ORANGE_FILL
        c = ws2.cell(row=row, column=3, value=page_totals['income'])
        c.number_format = MONEY_FORMAT
        c.font = Font(bold=True, color="008000", size=12)
        c.border = THIN_BORDER
        c.fill = ORANGE_FILL

        row += 2
        ws2.cell(row=row, column=1, value="数据来源: 银行清单各页底部算术合计，为权威数据").font = Font(italic=True, color="666666")

        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 16
        ws2.column_dimensions['C'].width = 16

    wb.save(filepath)
    print(f"  Excel: {filepath}")


# ============================================================
# HTML 报告
# ============================================================
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{title}</title>
<style>
  @page {{ size: A4; margin: 15mm; }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: "PingFang SC", "Microsoft YaHei", sans-serif; color: #333; padding: 20px; max-width: 900px; margin: 0 auto; font-size: 13px; line-height: 1.6; }}
  h1 {{ color: #1a237e; border-bottom: 3px solid #1a237e; padding-bottom: 8px; margin-bottom: 5px; font-size: 22px; }}
  h2 {{ color: #283593; margin-top: 25px; margin-bottom: 10px; font-size: 16px; border-left: 4px solid #283593; padding-left: 10px; }}
  .subtitle {{ color: #666; font-size: 13px; margin-bottom: 20px; }}
  .summary-box {{ background: linear-gradient(135deg, #e8eaf6, #c5cae9); border-radius: 8px; padding: 15px 20px; margin: 15px 0; }}
  .summary-box .big {{ font-size: 28px; font-weight: bold; color: #c62828; }}
  .summary-box .label {{ color: #555; font-size: 13px; }}
  .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(180px, 1fr)); gap: 10px; margin: 15px 0; }}
  .summary-card {{ background: #f5f5f5; border-radius: 6px; padding: 12px; text-align: center; }}
  .summary-card .value {{ font-size: 20px; font-weight: bold; }}
  .summary-card .value.expense {{ color: #c62828; }}
  .summary-card .value.income {{ color: #2e7d32; }}
  table {{ width: 100%; border-collapse: collapse; margin: 10px 0; font-size: 12px; }}
  th {{ background: #2F5496; color: white; padding: 8px 6px; text-align: center; font-size: 12px; }}
  td {{ padding: 6px; border: 1px solid #ddd; }}
  tr:nth-child(even) {{ background: #f9f9f9; }}
  tr.expense {{ background: #fff2f2; }}
  tr.income {{ background: #f2fff2; }}
  .amount {{ text-align: right; font-family: "Courier New", monospace; }}
  .amount.negative {{ color: #c62828; }}
  .amount.positive {{ color: #2e7d32; }}
  .highlight {{ background: #fff3e0; font-weight: bold; }}
  .footer {{ margin-top: 30px; padding-top: 10px; border-top: 1px solid #ccc; color: #999; font-size: 11px; text-align: center; }}
</style>
</head>
<body>
{content}
<div class="footer">
  报告生成时间: {timestamp} | 数据来源: {source}
</div>
</body>
</html>"""


def generate_wechat_html(analyzed):
    """生成微信支付 HTML 内容"""
    period = analyzed.get('period', {})
    period_str = f"{period.get('start', '')} - {period.get('end', '')}"
    holder = analyzed.get('holder', '')
    account = analyzed.get('account', '')
    days = analyzed.get('days', 1)

    content = f"""
<h1>微信支付交易明细分析报告</h1>
<div class="subtitle">时间段: {period_str} | 户名: {holder} ({account})</div>

<div class="summary-box">
  <div class="label">累计支出</div>
  <div class="big">&yen; {analyzed['expense_total']:,.2f}</div>
</div>

<div class="summary-grid">
  <div class="summary-card">
    <div class="label">支出笔数</div>
    <div class="value expense">{len(analyzed['expense_txs'])} 笔</div>
  </div>
  <div class="summary-card">
    <div class="label">收入合计</div>
    <div class="value income">&yen; {analyzed['income_total']:,.2f}</div>
  </div>
  <div class="summary-card">
    <div class="label">净支出</div>
    <div class="value expense">&yen; {analyzed['expense_total'] - analyzed['income_total']:,.2f}</div>
  </div>
  <div class="summary-card">
    <div class="label">日均支出({days}天)</div>
    <div class="value">&yen; {analyzed['expense_total']/max(days,1):,.2f}</div>
  </div>
</div>

<h2>按月统计</h2>
<table>
<tr><th>月份</th><th>支出笔数</th><th>支出金额</th><th>收入金额</th></tr>
"""
    total_exp = 0
    total_inc = 0
    for month, info in analyzed['by_month'].items():
        total_exp += info['expense']
        total_inc += info['income']
        content += f"""<tr>
  <td>{month}</td><td style="text-align:center">{info['count']}</td>
  <td class="amount negative">{info['expense']:,.2f}</td>
  <td class="amount positive">{info['income']:,.2f}</td>
</tr>"""

    content += f"""<tr class="highlight">
  <td><b>合计</b></td><td></td>
  <td class="amount negative"><b>{total_exp:,.2f}</b></td>
  <td class="amount positive"><b>{total_inc:,.2f}</b></td>
</tr></table>

<h2>按支付方式</h2>
<table>
<tr><th>支付方式</th><th>笔数</th><th>金额</th><th>占比</th></tr>
"""
    for method, info in sorted(analyzed['by_payment'].items(), key=lambda x: -x[1]['total']):
        pct = info['total'] / analyzed['expense_total'] * 100 if analyzed['expense_total'] > 0 else 0
        content += f"""<tr>
  <td>{method}</td><td style="text-align:center">{info['count']}</td>
  <td class="amount">{info['total']:,.2f}</td>
  <td style="text-align:center">{pct:.1f}%</td>
</tr>"""

    content += """</table>

<h2>支出金额分布</h2>
<table>
<tr><th>金额区间</th><th>笔数</th><th>占比</th></tr>
"""
    for rng, cnt in analyzed['ranges'].items():
        pct = cnt / len(analyzed['expense_txs']) * 100 if analyzed['expense_txs'] else 0
        bar = '█' * int(pct / 2)
        content += f"""<tr><td>{rng}</td><td style="text-align:center">{cnt}</td>
  <td>{pct:.1f}% {bar}</td></tr>"""

    content += """</table>

<h2>大额支出(>=500元)</h2>
<table>
<tr><th>序号</th><th>日期</th><th>金额</th><th>类型</th><th>交易对方</th></tr>
"""
    for idx, t in enumerate(analyzed['large_expenses'], 1):
        if t['amount'] < 500:
            break
        content += f"""<tr class="expense">
  <td style="text-align:center">{idx}</td><td>{t['date']}</td>
  <td class="amount negative">{t['amount']:,.2f}</td>
  <td>{t['type']}</td><td>{t['counterparty']}</td>
</tr>"""

    content += "</table>"
    return content


def generate_card_html(data):
    """生成银行卡 HTML 内容"""
    transactions = data['transactions']
    page_totals = data.get('page_totals')
    card_name = data.get('card_name', '银行卡')
    card_number = data.get('card_number', '')
    holder = data.get('holder', '')
    branch = data.get('branch', '')
    period = data.get('period', {})
    period_str = f"{period.get('start', '')} - {period.get('end', '')}"

    # 使用 page_totals 如果有，否则用逐笔合计
    if page_totals:
        display_expense = page_totals['expense']
        display_income = page_totals['income']
    else:
        display_expense = sum(abs(t['amount']) for t in transactions if t['amount'] < 0)
        display_income = sum(t['amount'] for t in transactions if t['amount'] > 0)

    content = f"""
<h1>{card_name}消费记录分析报告</h1>
<div class="subtitle">时间段: {period_str} | {branch} | 卡号: {card_number} | 户名: {holder}</div>

<div class="summary-box">
  <div class="label">累计支出{'(银行清单合计)' if page_totals else ''}</div>
  <div class="big">&yen; {display_expense:,.2f}</div>
</div>

<div class="summary-grid">
  <div class="summary-card">
    <div class="label">支出合计</div>
    <div class="value expense">&yen; {display_expense:,.2f}</div>
  </div>
  <div class="summary-card">
    <div class="label">收入合计</div>
    <div class="value income">&yen; {display_income:,.2f}</div>
  </div>
  <div class="summary-card">
    <div class="label">净支出</div>
    <div class="value expense">&yen; {display_expense - display_income:,.2f}</div>
  </div>
</div>
"""

    # 分页汇总
    if page_totals and page_totals.get('pages'):
        content += """
<h2>分页汇总(银行原始数据)</h2>
<table>
<tr><th>页码</th><th>支出(元)</th><th>收入(元)</th></tr>
"""
        for page in page_totals['pages']:
            content += f"""<tr>
  <td>{page['name']}</td>
  <td class="amount negative">{page['expense']:,.2f}</td>
  <td class="amount positive">{page['income']:,.2f}</td>
</tr>"""

        content += f"""<tr class="highlight">
  <td><b>合计</b></td>
  <td class="amount negative"><b>{page_totals['expense']:,.2f}</b></td>
  <td class="amount positive"><b>{page_totals['income']:,.2f}</b></td>
</tr></table>
"""

    # 交易明细
    content += """
<h2>交易明细</h2>
<table>
<tr><th>日期</th><th>金额(元)</th><th>类型</th><th>对方/说明</th></tr>
"""
    for tx in transactions:
        amount = tx['amount']
        cls = "expense" if amount < 0 else "income"
        amt_cls = "negative" if amount < 0 else "positive"
        content += f"""<tr class="{cls}">
  <td>{tx['date']}</td>
  <td class="amount {amt_cls}">{amount:,.2f}</td>
  <td>{tx.get('type', '')}</td><td>{tx.get('description', '')}</td>
</tr>"""

    content += "</table>"
    return content


def save_html(content, title, source, filepath):
    """保存 HTML 文件"""
    html = HTML_TEMPLATE.format(
        title=title,
        content=content,
        timestamp=datetime.now().strftime('%Y年%m月%d日 %H:%M'),
        source=source,
    )
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  HTML: {filepath}")
    return filepath


def html_to_pdf(html_path, pdf_path):
    """HTML 转 PDF (wkhtmltopdf 优先，Chrome headless 备选)"""
    try:
        result = subprocess.run(
            ['wkhtmltopdf', '--encoding', 'utf-8', '--page-size', 'A4',
             '--margin-top', '10mm', '--margin-bottom', '10mm',
             '--margin-left', '10mm', '--margin-right', '10mm',
             str(html_path), str(pdf_path)],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode == 0:
            print(f"  PDF:  {pdf_path}")
            return True
    except FileNotFoundError:
        pass

    try:
        chrome_paths = [
            '/Applications/Google Chrome.app/Contents/MacOS/Google Chrome',
            '/Applications/Chromium.app/Contents/MacOS/Chromium',
        ]
        for chrome in chrome_paths:
            if Path(chrome).exists():
                # URL 编码处理中文路径
                encoded_path = quote(str(html_path))
                result = subprocess.run(
                    [chrome, '--headless', '--disable-gpu',
                     f'--print-to-pdf={pdf_path}',
                     '--no-pdf-header-footer',
                     f'file://{encoded_path}'],
                    capture_output=True, text=True, timeout=30
                )
                if Path(pdf_path).exists():
                    print(f"  PDF:  {pdf_path}")
                    return True
    except Exception:
        pass

    print(f"  PDF:  跳过(需安装 wkhtmltopdf 或 Chrome)")
    return False


# ============================================================
# 主函数
# ============================================================
def main():
    parser = argparse.ArgumentParser(description='账单流水分析报告生成 (Excel + HTML + PDF)')
    parser.add_argument('--data', '-d', required=True, help='JSON 数据文件路径')
    parser.add_argument('--output-dir', '-o', default='.', help='输出目录(默认当前目录)')
    parser.add_argument('--name', '-n', help='报告文件名前缀(默认从JSON推断)')
    parser.add_argument('--formats', '-f', default='excel,html,pdf', help='输出格式,逗号分隔(默认: excel,html,pdf)')
    args = parser.parse_args()

    # 读取 JSON 数据
    with open(args.data, 'r', encoding='utf-8') as f:
        data = json.load(f)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    data_type = data.get('type', '')
    formats = set(args.formats.split(','))

    # 推断报告名称
    if args.name:
        report_name = args.name
    elif data_type == 'wechat':
        period = data.get('period', {})
        start = period.get('start', '').replace('-', '')
        end = period.get('end', '').replace('-', '')
        report_name = f"分析报告-微信支付({start}-{end})"
    elif data_type == 'bank_card':
        card_name = data.get('card_name', '银行卡')
        report_name = f"分析报告-{card_name}"
    else:
        report_name = "分析报告"

    print(f"\n处理: {report_name} (类型: {data_type})")
    print(f"  交易笔数: {len(data.get('transactions', []))}")

    if data_type == 'wechat':
        analyzed = analyze_wechat(data)
        print(f"  支出: {len(analyzed['expense_txs'])} 笔, {analyzed['expense_total']:,.2f} 元")
        print(f"  收入: {len(analyzed['income_txs'])} 笔, {analyzed['income_total']:,.2f} 元")

        if 'excel' in formats:
            create_wechat_excel(analyzed, output_dir / f"{report_name}.xlsx")

        if 'html' in formats or 'pdf' in formats:
            html_content = generate_wechat_html(analyzed)
            html_path = save_html(html_content, f"{report_name}",
                                  data.get('source', '微信支付交易明细'),
                                  output_dir / f"{report_name}.html")
            if 'pdf' in formats:
                html_to_pdf(html_path, output_dir / f"{report_name}.pdf")

    elif data_type == 'bank_card':
        page_totals = data.get('page_totals')
        if page_totals:
            print(f"  银行清单支出: {page_totals['expense']:,.2f} 元")
            print(f"  银行清单收入: {page_totals['income']:,.2f} 元")

        if 'excel' in formats:
            create_card_excel(data, output_dir / f"{report_name}.xlsx")

        if 'html' in formats or 'pdf' in formats:
            html_content = generate_card_html(data)
            html_path = save_html(html_content, f"{report_name}",
                                  data.get('source', '银行卡账户历史明细清单'),
                                  output_dir / f"{report_name}.html")
            if 'pdf' in formats:
                html_to_pdf(html_path, output_dir / f"{report_name}.pdf")

    else:
        print(f"  未知数据类型: {data_type}，跳过")
        return

    print(f"  完成! 输出目录: {output_dir}")


if __name__ == '__main__':
    main()
